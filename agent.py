#!/usr/bin/env python2
# -*- coding: utf-8 -*-

import openpyxl
import xlrd
import os
import os.path


def can_ignore(v):
    if v is None:
        return True
    if isinstance(v, str) or isinstance(v, unicode):
        v = v.strip()
        if len(v) == 0:
            return True
        if v == ".":
            return True
    return False

def is_excel(fn):
    _, ext = os.path.splitext(fn)
    if ext == ".xls" or ext == ".xlsx":
        if "~$" not in fn and "新建 Microsoft Excel 工作表" not in fn:
            return True
    return False


def row_col_letter(row, col):
    return "{}{}".format(openpyxl.utils.get_column_letter(col), row)

class AbsExcel:
    def __init__(self, fn):
        self.filename = fn
        if fn.endswith(".xlsx"):
            self.mode = 1
            self.wb = openpyxl.load_workbook(self.filename, read_only=True, data_only=True)
            self.ws = self.wb.active
        else:
            self.mode = 0
            self.wb = xlrd.open_workbook(self.filename, on_demand = True)
            self.ws = self.wb.sheet_by_index(0)


    def cell(self, row, col):
        if self.mode == 1:
            return self.ws.cell(row, col).value
        else:
            try:
                return self.ws.cell(row-1, col-1).value
            except Exception as e:
                return 0

    def get_sum(self, row, col):
        row_begin = 6
        result = 0.
        while row_begin < row:
            try:
                v = self.cell(row_begin, col)
                result += float(v)
            except Exception as e:
                if not can_ignore(v):
                    print "error: row={} col={}".format(row_begin, col)
                    if self.mode == 1:
                        print self.ws.cell(row, col)
                    else:
                        print self.ws.cell(row-1, col-1)
                    raise e

            row_begin += 1
        return result


    def get_init_value(self, col):
        row = 5
        v = self.cell(row, col)
        if v is None or v == "" or v == u"":
            return 0
        return v

    def close(self):
        if self.mode == 1:
            self.wb.close()
        else:
            self.wb.release_resources()

class Stat:
    def __init__(self, name, shoukuan = 0, fahuo = 0, jieyu = 0):
        self.name = name
        self.shoukuan = shoukuan
        self.fahuo = fahuo
        self.jieyu = jieyu

    def get_shoukuan(self):
        if self.shoukuan is None or self.shoukuan == "":
            return 0
        return self.shoukuan

    def get_fahuo(self):
        if self.fahuo is None or self.fahuo == "":
            return 0
        return self.fahuo

    def get_jieyu(self):
        if self.jieyu is None or self.jieyu == "":
            return 0
        return self.jieyu

    def __str__(self):
        return "[{}:{} {} {}]".format(self.name, self.shoukuan, self.fahuo, self.jieyu)

checklists = [
    # row col expected
    (2, 1, u"系统编码"),
    (2, 3, u"店名"),
    (2, 6, u"姓名"),
    (2, 9, u"电话号码"),
    (2, 12, u"地址"),
    (3, 5, u"货款"),
    (3, 8, u"定位费"),
    (3, 11, u"零售价任选"),
    (3, 14, u"套盒任选"),
    (3, 17, u"特殊政策"),
    (3, 20, u"套盒奖励"),
    (3, 23, u"保证金", False),
    (4, 1, u"日期"),
    (4, 2, u"凭证码"),
    (4, 3, u"方案"),
    (4, 4, u"备注"),
    (4, 5, u"收款"),
    (4, 6, u"发货"),
    (4, 7, u"结余"),
    (4, 8, u"收款"),
    (4, 9, u"发货"),
    (4, 10, u"结余"),
    (4, 11, u"收款"),
    (4, 12, u"发货"),
    (4, 13, u"结余"),
    (4, 14, u"收款"),
    (4, 15, u"发货"),
    (4, 16, u"结余"),
    (4, 17, u"收款"),
    (4, 18, u"发货"),
    (4, 19, u"结余"),
    (4, 20, u"收款"),
    (4, 21, u"发货"),
    (4, 22, u"结余"),
    (4, 23, u"收款", False),
    (4, 24, u"发货", False),
    (4, 25, u"结余", False),
    (5, 1, u"上期余额")
]

alignment = openpyxl.styles.Alignment(horizontal="center")
border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                right=openpyxl.styles.Side(border_style='thin', color='000000'),
                top=openpyxl.styles.Side(border_style='thin', color='000000'),
                bottom=openpyxl.styles.Side(border_style='thin', color='000000'))

class Agent:
    def __init__(self, fn):
        self.filename = fn
        self.is_valid = is_excel(fn)
        self.title = ""
        self.xitongbianma = ""
        self.dianming = ""
        self.xingming = ""
        self.dianhuahaoma = ""
        self.dizhi = ""
        self.huokuan = Stat("huokuan")
        self.dingweifei = Stat("dingweifei")
        self.lingshoujiarenxuan = Stat("lingshoujiarenxuan")
        self.taoherenxuan = Stat("taoherenxuan")
        self.teshuzhengce = Stat("teshuzhengce")
        self.taohejiangli = Stat("taohejiangli")
        self.baozhengjin = Stat("baozhengjin")
        self.excel = None
        self.error_msg = ""

    def parse(self):
        if not self.is_valid:
            return False
        try:
            self.excel = AbsExcel(self.filename)
        except Exception as e:
            print e
            return False

        #check format
        for checklist in checklists:
            row, col, content = checklist[0], checklist[1], checklist[2]
            if len(checklist) == 4 and checklist[3] == False:
                continue
            if self.excel.cell(row, col) != content:
                self.error_msg = "row:{} col:{} should be {}".format(row, col, content.encode("utf-8"))
                self.close()
                return False


        #get metadata
        self.title = self.excel.cell(1,1)
        self.xitongbianma = self.excel.cell(2,2)
        self.dianming = self.excel.cell(2,4)
        self.xingming = self.excel.cell(2,7)
        self.dianhuahaoma = self.excel.cell(2,10)
        self.dizhi = self.excel.cell(2,13)
        #get heji row number

        row = 0
        for row_index in xrange(6, 256):
            if self.excel.cell(row_index, 1) == u"合计":
                row = row_index
                break

        if row == 0:
            self.error_msg = "cannot find heji"
            self.close()
            return False

        # get last stat
        # self.huokuan.shoukuan = self.excel.cell(row, 5)
        # self.huokuan.fahuo = self.excel.cell(row, 6)
        # self.huokuan.jieyu = self.excel.cell(row, 7)
        # self.dingweifei.shoukuan = self.excel.cell(row, 8)
        # self.dingweifei.fahuo = self.excel.cell(row, 9)
        # self.dingweifei.jieyu = self.excel.cell(row, 10)
        # self.lingshoujiarenxuan.shoukuan = self.excel.cell(row, 11)
        # self.lingshoujiarenxuan.fahuo = self.excel.cell(row, 12)
        # self.lingshoujiarenxuan.jieyu = self.excel.cell(row, 13)
        # self.taoherenxuan.shoukuan = self.excel.cell(row, 14)
        # self.taoherenxuan.fahuo = self.excel.cell(row, 15)
        # self.taoherenxuan.jieyu = self.excel.cell(row, 16)
        # self.teshuzhengce.shoukuan = self.excel.cell(row, 17)
        # self.teshuzhengce.fahuo = self.excel.cell(row, 18)
        # self.teshuzhengce.jieyu = self.excel.cell(row, 19)
        # self.taohejiangli.shoukuan = self.excel.cell(row, 20)
        # self.taohejiangli.fahuo = self.excel.cell(row, 21)
        # self.taohejiangli.jieyu = self.excel.cell(row, 22)
        # self.baozhengjin.shoukuan = self.excel.cell(row, 23)
        # self.baozhengjin.fahuo = self.excel.cell(row, 24)
        # self.baozhengjin.jieyu = self.excel.cell(row, 25)

        # do compute from the beginning

        stat_map = {
            "huokuan": 5,
            "dingweifei": 8,
            "lingshoujiarenxuan": 11,
            "taoherenxuan": 14,
            "teshuzhengce": 17,
            "taohejiangli": 20,
            "baozhengjin": 23
        }

        try:
            for key, col in stat_map.iteritems():
                stat = self.__dict__[key]
                stat.shoukuan = self.excel.get_sum(row, col)
                stat.fahuo = self.excel.get_sum(row, col + 1)
                stat.jieyu = self.excel.get_init_value(col + 2) + stat.shoukuan - stat.fahuo
        except Exception as e:
            self.close()
            print "error"
            return False

        self.close()
        return True

    def close(self):
        if self.excel is not None:
            self.excel.close()

    def newyear(self, outfn):
        wb = openpyxl.Workbook()
        ws = wb.active
        #title
        title = self.title.replace("2021", "2022")
        ws.cell(1, 1, title)
        ws.cell(1, 1).alignment = alignment
        #format
        for checklist in checklists:
            row, col, content = checklist[0], checklist[1], checklist[2]
            ws.cell(row, col, content)
        #meta
        ws.cell(2,2,self.xitongbianma)
        ws.cell(2,4,self.dianming)
        ws.cell(2,7,self.xingming)
        ws.cell(2,10,self.dianhuahaoma)
        ws.cell(2,13,self.dizhi)
        #last
        ws.cell(5,7,self.huokuan.get_jieyu())
        ws.cell(5,10,self.dingweifei.get_jieyu())
        ws.cell(5,13,self.lingshoujiarenxuan.get_jieyu())
        ws.cell(5,16,self.taoherenxuan.get_jieyu())
        ws.cell(5,19,self.teshuzhengce.get_jieyu())
        # ws.cell(5,22,self.taohejiangli.jieyu)
        # ws.cell(5,25,self.baozhengjin.jieyu)
        ws.cell(5,22,0)
        ws.cell(5,25,0)
        #formulae
        for row in xrange(6, 50):
            for col in xrange(7, 28, 3):
                formula = "={}+{}-{}".format(row_col_letter(row-1,col), row_col_letter(row,col-2),row_col_letter(row,col-1))
                ws.cell(row,col,formula)
        row = 50
        ws.cell(row, 1, u"合计")
        for col in xrange(5, 26):
            if col % 3 != 1:
                formula = "=SUM({}:{})".format(row_col_letter(6, col), row_col_letter(row-1, col))
            else:
                formula = "={}".format(row_col_letter(row-1,col))
            ws.cell(row, col, formula)
        # merge
        ws.merge_cells("A1:Y1")
        ws.merge_cells("D2:E2")
        ws.merge_cells("G2:H2")
        ws.merge_cells("J2:K2")
        ws.merge_cells("M2:Y2")
        ws.merge_cells("A3:D3")
        ws.merge_cells("E3:G3")
        ws.merge_cells("H3:J3")
        ws.merge_cells("K3:M3")
        ws.merge_cells("N3:P3")
        ws.merge_cells("Q3:S3")
        ws.merge_cells("T3:V3")
        ws.merge_cells("W3:Y3")
        for col in xrange(5, 26, 3):
            ws.cell(3,col).alignment = alignment

        cell_range = "A1:Y50"
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border
        if self.xingming != "" and self.xingming is not None:
            ws.title = self.xingming
        else:
            ws.title = os.path.basename(self.filename).split(".")[0].decode("utf-8")

        wb.save(outfn)


    def get_month_info(self, root_dir):
        if not self.filename.startswith(root_dir):
            self.error_msg = "root_dir:{} not in filename:{}".format(root_dir, self.filename)
            return []

        info = os.path.dirname(self.filename).lstrip(root_dir).lstrip(os.path.sep).split(os.path.sep)

        if self.xingming != "" and self.xingming is not None:
            info.append(self.xingming.encode("utf-8"))
        else:
            info.append(os.path.basename(self.filename).split(".")[0])

        for stat in [self.huokuan, self.dingweifei, self.lingshoujiarenxuan, self.taoherenxuan, self.teshuzhengce, self.taohejiangli, self.baozhengjin]:
            info.append(stat.get_shoukuan())
            info.append(stat.get_fahuo())
            info.append(stat.get_jieyu())
        return info


if __name__ == "__main__":
    fn = u"/Users/fanan/Downloads/dashifu//2020客户明细表/江苏区域/苏南/无锡/仰双岱.xlsx"
    a = Agent(fn)
    assert a.is_valid
    assert a.parse()
    # for info in a.get_month_info(os.path.dirname(os.path.expanduser("~/Downloads/dashifu/"))):
    #     print info
    a.newyear("hello.xlsx")
